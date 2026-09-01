#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт пересборки index.html из трёх файлов-источников:
  source_data/classifier.rtf              — классификатор ОКПД2 (экспорт из
                                             Гаранта/Консультанта, формат RTF)
  source_data/national_regime_1875.xlsx   — таблица «код -> требование» по
                                             постановлению №1875 (лист "ПП 1875"
                                             или просто первый лист файла)
  source_data/okpd719_balls.xlsx          — таблица баллов по постановлению
                                             №719 (результат вашего парсера
                                             okpd719_points_parser.py)

Как обновить данные:
  1. Замените нужный файл в папке source_data/ на новую версию (имя файла
     менять не нужно, только содержимое).
  2. Запустите rebuild_static_data.bat (двойной клик).
  3. Скрипт перезапишет index.html и сам отправит изменения на GitHub.

Ничего в папке data/ (данные реестра) этот скрипт не трогает — тот
контур обновляется отдельным скриптом update_registry.py/.bat.
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("Не найден пакет openpyxl. Установите: python -m pip install openpyxl")
    sys.exit(1)

try:
    from striprtf.striprtf import rtf_to_text
except ImportError:
    print("Не найден пакет striprtf. Установите: python -m pip install striprtf")
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(ROOT, "source_data")
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template.html")
OUTPUT_PATH = os.path.join(ROOT, "index.html")

CLASSIFIER_PATH = os.path.join(SOURCE_DIR, "classifier.rtf")
REG1875_PATH = os.path.join(SOURCE_DIR, "national_regime_1875.xlsx")
REG719_PATH = os.path.join(SOURCE_DIR, "okpd719_balls.xlsx")

CODE_RE = re.compile(r"\d{2}(?:\.\d+)*")
FOOTNOTE_RE = re.compile(r"\*\(\d+\)")
EXCLUSION_RE = re.compile(r"\(\s*за\s+исключением([^)]*)\)", re.IGNORECASE)


def extract_codes_and_exclusions(code_raw):
    """Разбирает исходную строку кода на:
    - match_codes — коды, к которым реально относится строка,
    - exclude_codes — коды, которые явно ИСКЛЮЧЕНЫ формулировкой
      «(за исключением ...)» и не должны попадать под требование этой строки.

    Также убирает сноски вида «*(11)» — это ссылка на примечание, а не
    отдельный код (раньше по ошибке распознавался как код «11»)."""
    text = code_raw
    exclude_codes = []
    m = EXCLUSION_RE.search(text)
    if m:
        exclude_codes = CODE_RE.findall(m.group(1))
        text = EXCLUSION_RE.sub("", text)
    text = FOOTNOTE_RE.sub("", text)
    match_codes = CODE_RE.findall(text)
    return match_codes, exclude_codes


def log(msg):
    print(msg, flush=True)


def parse_classifier(path):
    """RTF-экспорт классификатора ОКПД2 -> список [код, наименование]."""
    log(f"Читаю классификатор: {path}")
    raw = open(path, encoding="cp1251", errors="replace").read()
    text = rtf_to_text(raw)
    pattern = re.compile(r"^(\d{2}(?:\.\d+)*)\|(.*)$")
    classifier = {}
    for line in text.split("\n"):
        line = line.rstrip()
        m = pattern.match(line)
        if not m:
            continue
        code, name = m.group(1), m.group(2)
        name = name.rstrip("|").strip()
        if name:
            classifier[code] = name
    arr = [[k, v] for k, v in classifier.items()]
    log(f"  найдено кодов: {len(arr)}")
    if len(arr) < 1000:
        log("  ВНИМАНИЕ: подозрительно мало кодов — проверьте, тот ли это файл "
            "и не изменилась ли его внутренняя структура.")
    return arr


def get_sheet(wb, preferred_name):
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    log(f"  лист «{preferred_name}» не найден, беру первый лист файла ({wb.sheetnames[0]})")
    return wb.worksheets[0]


def parse_1875(path):
    log(f"Читаю таблицу постановления №1875: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = get_sheet(wb, "ПП 1875")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        prilozhenie, pozicia, code, name, trebovanie = (list(row) + [None] * 5)[:5]
        code = str(code).strip()
        rows.append({
            "prilozhenie": prilozhenie,
            "pozicia": pozicia,
            "code": code,
            "name": (name or "").strip(),
            "trebovanie": (trebovanie or "").strip(),
        })
    log(f"  строк: {len(rows)}")
    return rows


def parse_719(path):
    log(f"Читаю таблицу баллов по постановлению №719: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = get_sheet(wb, "Баллы по кодам ОКПД2")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        (razdel, code_raw, name_code, name_note, trebovaniya,
         note_num, perehod, rule, status) = (list(row) + [None] * 9)[:9]
        code_raw = str(code_raw).strip()
        match_codes, exclude_codes = extract_codes_and_exclusions(code_raw)
        rows.append({
            "razdel": razdel,
            "code_display": code_raw,
            "match_codes": match_codes,
            "exclude_codes": exclude_codes,
            "name_code": (name_code or "").strip(),
            "name_note": (name_note or "").strip(),
            "trebovaniya": (trebovaniya or "").replace("\xa0", " ").strip(),
            "note_num": note_num,
            "perehod": (perehod or "").replace("\xa0", " ").strip(),
            "rule": rule,
            "status": status,
        })
    log(f"  строк: {len(rows)}")
    return rows


def esc(s):
    return s.replace("</script", "<\\/script")


def build_html(classifier_arr, reg1875, reg719):
    log(f"Собираю index.html из шаблона: {TEMPLATE_PATH}")
    template = open(TEMPLATE_PATH, encoding="utf-8").read()

    out = (template
           .replace("__CLASSIFIER_JSON__", esc(json.dumps(classifier_arr, ensure_ascii=False, separators=(",", ":"))))
           .replace("__REG1875_JSON__", esc(json.dumps(reg1875, ensure_ascii=False, separators=(",", ":"))))
           .replace("__REG719_JSON__", esc(json.dumps(reg719, ensure_ascii=False, separators=(",", ":")))))

    if "__CLASSIFIER_JSON__" in out or "__REG1875_JSON__" in out or "__REG719_JSON__" in out:
        raise RuntimeError("В шаблоне остались незаполненные метки — что-то не так с template.html")

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(out)
    log(f"Готово: {OUTPUT_PATH} ({os.path.getsize(OUTPUT_PATH) / 1024 / 1024:.2f} МБ)")


def main():
    missing = [p for p in (CLASSIFIER_PATH, REG1875_PATH, REG719_PATH) if not os.path.isfile(p)]
    if missing:
        for p in missing:
            log(f"Не найден файл-источник: {p}")
        raise RuntimeError(
            "Не хватает файлов-источников в папке source_data/. "
            "Ничего не пересобрано, index.html не изменён."
        )

    classifier_arr = parse_classifier(CLASSIFIER_PATH)
    reg1875 = parse_1875(REG1875_PATH)
    reg719 = parse_719(REG719_PATH)
    build_html(classifier_arr, reg1875, reg719)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"ОШИБКА: {e}")
        sys.exit(1)
